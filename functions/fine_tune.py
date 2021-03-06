from __future__ import print_function, division

import torch
import torch.nn as nn
import torch.optim as optim
import numpy as np
import matplotlib.pyplot as plt
import time
import copy
import openpyxl
import math

'''
TODOS:
- Check the code for binomial and poisson
- Make sure that model output covers everything for cheng, poisson, binomial, fix_a, learn-a and weighted_softmax
- Check the code for weighted_softmax
'''

def exp_lr_scheduler(optimizer, epoch, init_lr=0.001, lr_decay_epoch=7):
    """Decay learning rate by a factor of 0.1 every lr_decay_epoch epochs."""
    lr = init_lr * (0.1 ** (epoch // lr_decay_epoch))

    if epoch % lr_decay_epoch == 0:
        print('LR is set to {}'.format(lr))

    for param_group in optimizer.param_groups:
        param_group['lr'] = lr

    return optimizer


def visualize_model(model, num_images=6):
    images_so_far = 0
    fig = plt.figure()

    for i, data in enumerate(dset_loaders['val']):
        inputs, labels = data
        if use_gpu:
            inputs, labels = Variable(inputs.cuda()), Variable(labels.cuda())
        else:
            inputs, labels = Variable(inputs), Variable(labels)

        outputs = model(inputs)
        _, preds = torch.max(outputs.data, 1)

        for j in range(inputs.size()[0]):
            images_so_far += 1
            ax = plt.subplot(num_images // 2, 2, images_so_far)
            ax.axis('off')
            ax.set_title('predicted: {}'.format(dset_classes[labels.data[j]]))
            imshow(inputs.cpu().data[j])

            if images_so_far == num_images:
                return

def train_model(model, optim_str, lr_scheduler, dset_loaders, \
                dset_sizes, writer, use_gpu=True, num_epochs=25, batch_size=4, num_log=100, \
                init_lr=0.001, lr_decay_epoch=7, regression=False, learn_a=False,
                cross_loss=1., multi_loss=0., write_log=False,
                numOut=6, logname='logs.xlsx', iter_loc=12,
                multi_coeff=[1, 1, 1], single_coeff=[1, 1, 1], KL=False,
                poisson=False, binomial=False, cheng=False, algo=None,
                mae_loss=False, weighted_softmax=False, test=False,
               momentum = 0, weight_decay = 0, fix_a = False, cheng_lambda = 0,
               weighted_softmax_2 = False, softmax_matrices = []):


    if algo is 'KL':
        KL = True
        cross_loss = 1.
        multi_loss = 0.
    elif algo is 'softmax':
        cross_loss = 1.
        multi_loss = 0.
    elif algo is 'sigmoid':
        cross_loss = 0.
        multi_loss = 1.
    elif algo is 'learn_a':
        learn_a = True
        cross_loss = 1.
        multi_loss = 0.
    elif algo is 'fix_a':
        fix_a = True
        cross_loss = 1.
        multi_loss = 0.
    elif algo is 'learn_a_mae':
        learn_a = True
        cross_loss = 1.
        multi_loss = 0.
        mae_loss = True
    elif algo is 'fix_a_mae':
        fix_a = True
        cross_loss = 1.
        multi_loss = 0.
        mae_loss = True
    elif algo is 'poisson':
        KL = True
        poisson = True
        cross_loss = 0.
        multi_loss = 0.
        multi_coeff = [1]
        single_coeff = [1]
    elif algo is 'binomial':
        KL = True
        binomial = True
        cross_loss = 0.
        multi_loss = 0.
        multi_coeff = [1]
        single_coeff = [1]
    elif algo is 'cheng':
        print('Cheng')
        KL = True
        cheng = True
        cross_loss = 0.
        multi_loss = 0.
    elif algo is 'weighted_softmax_2':
        weighted_softmax_2 = True
        KL = True
    elif algo is 'weighted_softmax':
        weighted_softmax = True
        KL = True
    elif algo is 'regression':
        regression = True
        cross_loss = 1.
        multi_loss = 0.
    elif algo is 'regression_mae':
        regression = True
        cross_loss = 1.
        multi_loss = 0.
        mae_loss = True


    device = torch.device("cuda" if use_gpu else "cpu")
    result_log = []
    since = time.time()

    last_model = model
    best_model = model
    best_rmse = 100.0
    

        
    if learn_a or fix_a:
        model.fc = nn.Sequential(model.fc, nn.Softmax(dim=1), nn.Linear(model.fc.out_features, 1, bias = False)).to(device)
        if fix_a:
            model.fc[2].weight = torch.nn.Parameter(torch.from_numpy(np.arange(0, numOut)).type(torch.FloatTensor).view(1, -1).to(device), False)

    if poisson or binomial or regression:
        model.fc = nn.Linear(model.fc.in_features, 1).to(device)
        #learn_a = True

    if cheng:
        for k in range(multi_coeff.shape[0]):
            temp = multi_coeff[k,:]
            temp= temp/np.sum(temp)
            temp = np.cumsum(temp[::-1])
            multi_coeff[k,:] = temp[::-1]
        diff_arr = np.eye(numOut)
        for k in range(numOut-1):
            diff_arr[k+1, k] = -1
        diff_arr = torch.from_numpy(diff_arr).type(torch.FloatTensor).to(device)
        #print('Loss is cheng, multi_coeff is ' + str(multi_coeff))
            
    if poisson:
        log_j_fact = np.log(np.asarray([math.factorial(j) for j in range(numOut)]))
        ones_vec = torch.ones(numOut).type(torch.FloatTensor).view(1, numOut).to(device)
        j_vec = torch.arange(0, numOut).type(torch.FloatTensor).view(1, numOut).to(device)
        log_j_fact = torch.from_numpy(log_j_fact).type(torch.FloatTensor).view(1, numOut).to(device)

    if binomial:
        log_j_binom = np.log(np.asarray([math.factorial(numOut - 1) /
                                         (math.factorial(j) * math.factorial(numOut - 1 - j)) for j in range(numOut)]))

        ones_vec = torch.ones(numOut).type(torch.FloatTensor).view(1, numOut).to(device)
        j_vec = torch.arange(0, numOut).type(torch.FloatTensor).view(1, numOut).to(device)
        log_j_binom = torch.from_numpy(log_j_binom).type(torch.FloatTensor).view(1, numOut).to(device)

    if weighted_softmax:
        Accumulators = np.zeros((numOut, numOut * (numOut - 1)))
        for l in range(numOut):
            # A = np.zeros((numOut, numOut-1))
            for k in range(numOut - 1):
                Accumulators[np.maximum(0, l - k):np.minimum(numOut, l + k + 1), l * (numOut - 1) + k] = 1
        Accumulators = torch.from_numpy(Accumulators).type(torch.FloatTensor).to(device)

    if(optim_str=='adam'):
        optimizer = optim.Adam(filter(lambda p: p.requires_grad, model.parameters()), lr=init_lr, weight_decay=weight_decay)
    elif(optim_str=='sgd'):
        optimizer = optim.SGD(filter(lambda p: p.requires_grad, model.parameters()), lr=init_lr, momentum=momentum) #, weight_decay=weight_decay)
    #print('Model is  ' + str(model))
    #print('Requires grad is ' + str(model.reg.bias.requires_grad))
    for epoch in range(num_epochs):
        if (write_log):
            print('Epoch {}/{}'.format(epoch, num_epochs - 1))
            print('-' * 10)

        # Each epoch has a training and validation phase
        for phase in ['train', 'val']:
            if phase == 'train':
                batch_count = 0
                if lr_scheduler is not None:
                    optimizer = lr_scheduler(optimizer, epoch, init_lr=init_lr, lr_decay_epoch=lr_decay_epoch)
                model.train(True)  # Set model to training mode
            else:
                model.train(False)  # Set model to evaluate mode

            running_loss = 0.0
            running_corrects = 0
            running_cir1 = 0
            running_mse = 0.0
            running_mae = 0.0
            # Iterate over data.
            for data in dset_loaders[phase]:
                # get the inputs
                inputs, labels = data

                # wrap them in Variable
                inputs, labels = inputs.to(device), labels.to(device)

                # zero the parameter gradients
                optimizer.zero_grad()

                # forward
                outputs = model(inputs)
                
                #print('Model is ' + str(model))
                #print('Outputs size is ' + str(outputs.size()))

                # result_log.append((phase, epoch, labels.data.cpu().numpy(), outputs.data.cpu().numpy()))


                loss = torch.Tensor(1)
                loss = 0.0
                if learn_a or fix_a:
                    # print(labels)
                    labels = labels.type(torch.FloatTensor).to(device).view(-1,1)

                    if (mae_loss):
                        criterion = torch.nn.L1Loss()
                    else:
                        criterion = torch.nn.MSELoss()

                    if learn_a:
                        sigmoid_step = torch.nn.Sigmoid()
                        preds = (numOut - 1) * sigmoid_step(outputs)
                    else:
                        preds = outputs

                    loss += criterion(preds, labels)
                elif poisson:
                    if KL:
                        # print('KL div')
                        labels_multi = []
                        for label in labels.data:
                            extend = int((len(single_coeff) - 1) / 2)
                            label_multi = np.zeros(numOut + 2 * extend)
                            label_multi[label:label + 2 * extend + 1] = single_coeff
                            if extend is not 0:
                                label_multi = label_multi[extend:-extend]
                                label_multi = label_multi / np.sum(label_multi)
                                # print('KL divergence labels ' + str(label_multi))
                            labels_multi.append(label_multi)

                        labelsv = torch.FloatTensor(labels_multi).view(-1, numOut).to(device)



                    softplus_step = torch.nn.Softplus()
                    preds = softplus_step(outputs)

                    # print(preds.data.cpu().numpy())
                    outputs = torch.mm(preds, ones_vec)
                    outputs = j_vec * torch.log(outputs) - outputs - log_j_fact
                    # print(outputs)

                    log_soft = nn.LogSoftmax(dim = 1)
                    outputs_log_softmax = log_soft(outputs)

                    criterion = nn.KLDivLoss()

                    # print(outputs_softmax)
                    # print(labelsv)
                    loss = criterion(outputs_log_softmax, labelsv)

                elif binomial:
                    if KL:
                        # print('KL div')
                        labels_multi = []
                        for label in labels.data:
                            extend = int((len(single_coeff) - 1) / 2)
                            label_multi = np.zeros(numOut + 2 * extend)
                            label_multi[label:label + 2 * extend + 1] = single_coeff
                            if extend is not 0:
                                label_multi = label_multi[extend:-extend]
                                label_multi = label_multi / np.sum(label_multi)
                                # print('KL divergence labels ' + str(label_multi))
                            labels_multi.append(label_multi)

                        labelsv = torch.FloatTensor(labels_multi).view(-1, numOut).to(device)



                    sigmoid_step = torch.nn.Sigmoid()
                    preds = ((sigmoid_step(outputs) - 0.5) * 0.99) + .5
                    outputs = torch.mm(preds, ones_vec)
                    outputs = j_vec * torch.log(outputs) + (numOut - 1 - j_vec) * torch.log(1 - outputs) + log_j_binom

                    log_soft = nn.LogSoftmax(dim = 1)
                    outputs_log_softmax = log_soft(outputs)
                    criterion = nn.KLDivLoss()

                    # print(outputs_softmax)
                    # print(labelsv)
                    loss = criterion(outputs_log_softmax, labelsv)

                elif cheng:
                    labels_multi = []
                    for label in labels.data:
                        labels_multi.append(multi_coeff[label, :])
                    labelsv = torch.FloatTensor(labels_multi).view(-1, numOut).to(device)
                    criterion = nn.MultiLabelSoftMarginLoss()
                    loss += criterion(outputs, labelsv)

                    if cheng_lambda > 0:
                        sigmoid_step = torch.nn.Sigmoid()
                        cdf = sigmoid_step(outputs)
                        pmf = (cdf @ diff_arr) / .2
                        criterion = nn.CrossEntropyLoss()
                        loss += cheng_lambda * criterion(outputs, labels)

                elif weighted_softmax:
                    _, preds = torch.max(outputs.data, 1)
                    #print('Multi loss is ' + str(multi_loss))
                    if multi_loss > 0.:
                        if KL:
                            # print('KL div')
                            soft = nn.Softmax()
                            outputs_soft = soft(outputs)

                            labels_multi = []
                            for label in labels.data:
                                label_multi = np.zeros(numOut * (numOut - 1))
                                label_multi[label * (numOut - 1):(label + 1) * (numOut - 1)] = 1
                                labels_multi.append(label_multi)

                            labelsv = torch.FloatTensor(labels_multi).view(-1, numOut * (numOut - 1)).to(device)

                            outputs_accumulated_soft = torch.mm(outputs_soft, Accumulators)
                            outputs_log_softmax = torch.log(outputs_accumulated_soft)
                            criterion = nn.KLDivLoss()

                            # print('Outputs are ' + str(outputs_log_softmax.data.cpu().numpy()[10,:]))
                            # print('Labels are ' + str(labelsv.cpu().data.numpy()[10,:]))
                            loss += (multi_loss * (numOut-1) * criterion(outputs_log_softmax, labelsv))/3.0

                    if cross_loss > 0.:
                        _, preds = torch.max(outputs.data, 1)
                        if cross_loss > 0.:
                            if KL:
                                # print('KL div')
                                labels_multi = []
                                for label in labels.data:
                                    labels_multi.append(single_coeff[label, :])

                                labelsv = torch.FloatTensor(labels_multi).view(-1, numOut).to(device)

                                log_soft = nn.LogSoftmax(dim = 1)
                                outputs_log_softmax = log_soft(outputs)
                                criterion = nn.KLDivLoss()
                                loss += cross_loss * criterion(outputs_log_softmax, labelsv)

                            else:
                                criterion = nn.CrossEntropyLoss()
                                loss += cross_loss * criterion(outputs, labels)


                elif weighted_softmax_2:

                    #print('Multi loss is ' + str(multi_loss))
                    #print('Size of the outputs is ' + str(outputs.size()))
                    criterion = nn.CrossEntropyLoss()
                    _, preds = torch.max(outputs.data, 1)

                    for coeff, matrice in softmax_matrices:
                        weighted_out = torch.mm(outputs, matrice)
                        loss += criterion(weighted_out, labels)


                else:
                    _, preds = torch.max(outputs.data, 1)
                    if cross_loss > 0.:
                        if KL:
                            labels_multi = []
                            for label in labels.data:
                                labels_multi.append(single_coeff[label, :])

                            labelsv = torch.FloatTensor(labels_multi).view(-1, numOut).to(device)

                            log_soft = nn.LogSoftmax(dim = 1)
                            outputs_log_softmax = log_soft(outputs)
                            criterion = nn.KLDivLoss()
                            loss += cross_loss * criterion(outputs_log_softmax, labelsv)

                        else:
                            criterion = nn.CrossEntropyLoss()
                            loss += cross_loss * criterion(outputs, labels)

                    if multi_loss > 0.:
                        labels_multi = []
                        for label in labels.data:
                            labels_multi.append(multi_coeff[label, :])

                        labelsv = torch.FloatTensor(labels_multi).cuda().view(-1, numOut)
                        criterion = nn.MultiLabelSoftMarginLoss()
                        loss += multi_loss * criterion(outputs, labelsv)

                # backward + optimize only if in training phase
                if phase == 'train':
                    batch_count += 1
                    if (np.mod(batch_count, num_log) == 0):

                        batch_acc = float(running_corrects) / float(batch_count * batch_size)
                        batch_cir1 = float(running_cir1) / float(batch_count * batch_size)
                        batch_rmse = np.sqrt(float(running_mse) / float(batch_count * batch_size))
                        batch_mae = float(running_mae) / float(batch_count * batch_size)


                        if (write_log):

                            print('{}/{}, acc: {:.4f}, CIR-1: {:.4f}, RMSE: {:.4f}, MAE: {:.4f}'
                                  .format(batch_count, len(dset_loaders['train']),
                                          batch_acc, batch_cir1, batch_rmse, batch_mae))
                            '''
                            if(learn_a or fix_a):
                                print('Weights are ' + str(model.fc[1].weight) + ', bias is ' + str(model.fc[1].bias))
                            else: 
                                print('Weights are ' + str(model.fc.weight) + ', bias is ' + str(model.fc.bias))
                            '''
                        ''' TO BE DELETED '''
                        
                        if cheng_lambda > 0:
                            cdf_np = cdf[5,:].data.to('cpu').numpy()
                            print('CDF is ' + str(np.round(cdf_np, decimals = 2)))
                            pmf_np = pmf[5,:].data.to('cpu').numpy()
                            #print('PMF before softmax is ' + str(np.round(pmf_np, decimals = 2)))
                            pmf_np = np.exp(pmf_np)
                            pmf_np = pmf_np / np.sum(pmf_np)
                            print('PMF is ' + str(np.round(pmf_np, decimals = 2)))

                        ''' TO BE DELETED '''
                    loss.backward()
                    optimizer.step()

                # statistics
                running_loss += loss.item()

                if (learn_a or fix_a or poisson or binomial or cheng):
                    if cheng:
                        preds_numpy = (outputs.data.cpu().numpy() > 0.0).astype(np.int)
                        preds_numpy = (np.sum(np.cumprod(preds_numpy, axis=1), axis=1) - 1).reshape(-1, 1)
                    else:
                        preds_numpy = preds.data.cpu().numpy()
                        if binomial:
                            preds_numpy = preds_numpy * (numOut)

                        if poisson:
                            preds_numpy = np.floor(preds_numpy)
                        preds_numpy = np.round(preds_numpy)

                    preds_numpy[preds_numpy < 0] = 0
                    preds_numpy[preds_numpy > numOut - 1] = numOut - 1
                    labels_numpy = labels.data.cpu().numpy().reshape(-1, 1)

                    running_cir1 += np.sum(np.abs(preds_numpy - labels_numpy) <= 1)
                    running_mse += np.sum((preds_numpy - labels_numpy) * (preds_numpy - labels_numpy))
                    running_mae += np.sum(np.abs(preds_numpy - labels_numpy))
                    running_corrects += np.sum(np.abs(preds_numpy - labels_numpy) < .3)
                else:
                    running_cir1 += torch.sum(torch.abs(preds - labels.data) <= 1)
                    running_corrects += torch.sum(preds == labels.data)
                    running_mse += torch.sum((preds - labels.data) * (preds - labels.data))
                    running_mae += torch.sum(torch.abs(preds - labels.data))

            epoch_loss = float(running_loss) / float(dset_sizes[phase])
            epoch_acc = float(running_corrects) / float(dset_sizes[phase])
            epoch_cir1 = float(running_cir1) / float(dset_sizes[phase])
            epoch_rmse = np.sqrt((float(running_mse) / float(dset_sizes[phase])))
            epoch_mae = float(running_mae) / float(dset_sizes[phase])

            writer.add_scalar(phase + ' loss', epoch_loss, epoch)
            writer.add_scalar(phase + ' accuracy', epoch_acc, epoch)
            writer.add_scalar(phase + ' CIR-1', epoch_cir1, epoch)
            writer.add_scalar(phase + 'RMSE', epoch_rmse, epoch)
            writer.add_scalar(phase + 'MAE', epoch_mae, epoch)
            if phase == 'train':
                epoch_loss_tr = epoch_loss
                epoch_acc_tr = epoch_acc
                epoch_rmse_tr = epoch_rmse
                epoch_mae_tr = epoch_mae
                epoch_cir1_tr = epoch_cir1
            if (write_log):
                print('{} Loss: {:.4f} Acc: {:.4f} CIR-1: {:.4f} RMSE {:.4f} MAE {:.4f}'.format(
                    phase, epoch_loss * 1000, epoch_acc, epoch_cir1, epoch_rmse, epoch_mae))

            book = openpyxl.load_workbook(logname)
            sheet = book.active
            current_row = sheet.max_row

            sheet.cell(row=current_row, column=iter_loc + 11).value = epoch + 1
            sheet.cell(row=current_row, column=iter_loc + 12).value = epoch_acc_tr
            sheet.cell(row=current_row, column=iter_loc + 13).value = epoch_acc
            sheet.cell(row=current_row, column=iter_loc + 14).value = epoch_rmse_tr
            sheet.cell(row=current_row, column=iter_loc + 15).value = epoch_rmse
            sheet.cell(row=current_row, column=iter_loc + 16).value = epoch_mae_tr
            sheet.cell(row=current_row, column=iter_loc + 17).value = epoch_mae
            sheet.cell(row=current_row, column=iter_loc + 18).value = epoch_cir1_tr
            sheet.cell(row=current_row, column=iter_loc + 19).value = epoch_cir1
            book.save(logname)
            last_model = copy.deepcopy(model)

            # deep copy the model
            if phase == 'val' and epoch_rmse < best_rmse:
                best_rmse = epoch_rmse
                best_model = copy.deepcopy(model)
                sheet.cell(row=current_row, column=iter_loc).value = epoch + 1
                sheet.cell(row=current_row, column=iter_loc + 1).value = epoch_loss_tr
                sheet.cell(row=current_row, column=iter_loc + 2).value = epoch_loss
                sheet.cell(row=current_row, column=iter_loc + 3).value = epoch_acc_tr
                sheet.cell(row=current_row, column=iter_loc + 4).value = epoch_acc
                sheet.cell(row=current_row, column=iter_loc + 5).value = epoch_rmse_tr
                sheet.cell(row=current_row, column=iter_loc + 6).value = epoch_rmse
                sheet.cell(row=current_row, column=iter_loc + 7).value = epoch_mae_tr
                sheet.cell(row=current_row, column=iter_loc + 8).value = epoch_mae
                sheet.cell(row=current_row, column=iter_loc + 9).value = epoch_cir1_tr
                sheet.cell(row=current_row, column=iter_loc + 10).value = epoch_cir1
                book.save(logname)
        if (write_log):
            print()

    time_elapsed = time.time() - since
    print('Training complete in {:.0f}m {:.0f}s'.format(
        time_elapsed // 60, time_elapsed % 60))
    print('Best val RMSE: {:4f}'.format(best_rmse))

    return best_model, last_model, result_log


def make_weights_for_balanced_classes(images, nclasses):
    '''
    Creates a weight vector which can be used in a weighted sampler functionfor creating balanced batches.

    :param images: Images taken from an ImageFolder object
    :param nclasses: Number of classes

    :return:
    :weight: Vector for weights
    :weight_per_class: Weights for classes
    '''

    count = [0] * nclasses
    for item in images:
        count[item[1]] += 1
    weight_per_class = [0.] * nclasses
    N = float(sum(count))
    for i in range(nclasses):
        weight_per_class[i] = N / float(count[i])
    weight = [0] * len(images)
    for idx, val in enumerate(images):
        weight[idx] = weight_per_class[val[1]]
    return weight, weight_per_class


def imshow(inp, title=None):
    """Imshow for Tensor."""
    inp = inp.numpy().transpose((1, 2, 0))
    mean = np.array([0.485, 0.456, 0.406])
    std = np.array([0.229, 0.224, 0.225])
    inp = std * inp + mean
    plt.imshow(inp)
    if title is not None:
        plt.title(title)
    plt.pause(0.001)  # pause a bit so that plots are updated


def train_model_balanced(model, criterion, optimizer, lr_scheduler, dset_loaders, \
                         dset_sizes, writer, use_gpu=True, num_epochs=25, batch_size=4, \
                         num_train=100, num_test=10, init_lr=0.001, lr_decay_epoch=7, \
                         multilabel=False, multi_prob=False, logname='logs.xlsx', iter_loc=12):
    '''
    Obsolete code, has to be rewritten if needed
    '''
    since = time.time()

    best_model = model
    best_cir1 = 0.0

    book = openpyxl.load_workbook(logname)
    sheet = book.active
    current_row = sheet.max_row
    for epoch in range(num_epochs):
        print('Epoch {}/{}'.format(epoch, num_epochs - 1))
        print('-' * 10)
        # Iterate over data.
        batch_count = 0
        #
        if lr_scheduler is not None:
            optimizer = lr_scheduler(optimizer, epoch, init_lr=init_lr, lr_decay_epoch=lr_decay_epoch)
        model.train(True)  # Set model to training mode
        for opt_iter in range(num_test):
            running_loss = 0.0
            running_corrects = 0
            running_cir1 = 0

            for k in range(num_train):
                # get the inputs
                inputs, labels = next(iter(dset_loaders['train']))

                # wrap them in Variable
                if use_gpu:
                    inputs, labels = Variable(inputs.cuda()), \
                                     Variable(labels.cuda())
                else:
                    inputs, labels = Variable(inputs), Variable(labels)

                # zero the parameter gradients
                optimizer.zero_grad()

                # forward
                outputs = model(inputs)
                _, preds = torch.max(outputs.data, 1)
                if (not multilabel):
                    loss = criterion(outputs, labels)
                else:
                    labels_multi = []
                    for label in labels.data:
                        label_multi = np.zeros(11)

                        if (multi_prob):
                            label_multi[label] = .5
                            label_multi[label + 1] = 1
                            label_multi[label + 2] = .5
                        else:
                            label_multi[label:label + 3] = 1

                        label_multi = label_multi[1:-1]
                        labels_multi.append(label_multi)
                    labelsv = Variable(torch.FloatTensor(labels_multi).cuda()).view(-1, 9)
                    loss = criterion(outputs, labelsv)

                # backward + optimize only if in training phase
                loss.backward()
                optimizer.step()

                # statistics
                running_loss += loss.item()
                running_corrects += torch.sum(preds == labels.data)
                running_cir1 += torch.sum(torch.abs(preds - labels.data) <= 1)

            epoch_loss = running_loss / (num_train * batch_size)
            epoch_acc = running_corrects / (num_train * batch_size)
            epoch_cir1 = running_cir1 / (num_train * batch_size)
            epoch_loss_tr = epoch_loss
            epoch_cir1_tr = epoch_cir1
            writer.add_scalar('train loss', epoch_loss, epoch)
            writer.add_scalar('train accuracy', epoch_acc, epoch)
            writer.add_scalar('train CIR-1', epoch_cir1, epoch)

            print('{}/{}, Loss: {:.4f} Acc: {:.4f} CIR-1: {:.4f}'
                  .format(opt_iter + 1, num_test, epoch_loss, epoch_acc, epoch_cir1))

            # deep copy the model
        model.train(False)  # Set model to evaluate mode
        running_loss = 0.0
        running_corrects = 0
        running_cir1 = 0

        for data in dset_loaders['val']:
            # get the inputs
            inputs, labels = data

            # wrap them in Variable
            if use_gpu:
                inputs, labels = Variable(inputs.cuda()), \
                                 Variable(labels.cuda())
            else:
                inputs, labels = Variable(inputs), Variable(labels)

            # zero the parameter gradients
            optimizer.zero_grad()

            # forward
            outputs = model(inputs)
            _, preds = torch.max(outputs.data, 1)
            if (not multilabel):
                loss = criterion(outputs, labels)
            else:
                labels_multi = []
                for label in labels.data:
                    label_multi = np.zeros(11)

                    if (multi_prob):
                        label_multi[label] = .5
                        label_multi[label + 1] = 1
                        label_multi[label + 2] = .5
                    else:
                        label_multi[label:label + 3] = 1

                    label_multi = label_multi[1:-1]
                    labels_multi.append(label_multi)
                labelsv = Variable(torch.FloatTensor(labels_multi).cuda()).view(-1, 9)
                loss = criterion(outputs, labelsv)

            # statistics
            running_loss += loss.item()
            running_corrects += torch.sum(preds == labels.data)
            running_cir1 += torch.sum(torch.abs(preds - labels.data) <= 1)

        epoch_loss = running_loss / dset_sizes['val']
        epoch_acc = running_corrects / dset_sizes['val']
        epoch_cir1 = running_cir1 / dset_sizes['val']
        writer.add_scalar('val loss', epoch_loss, epoch)
        writer.add_scalar('val accuracy', epoch_acc, epoch)
        writer.add_scalar('val CIR-1', epoch_cir1, epoch)

        print('Val Loss: {:.4f} Acc: {:.4f} CIR-1: {:.4f}'.format(epoch_loss, epoch_acc, epoch_cir1))

        if epoch_cir1 > best_cir1:
            best_cir1 = epoch_cir1
            best_model = copy.deepcopy(model)
            sheet.cell(row=current_row, column=iter_loc).value = epoch + 1
            sheet.cell(row=current_row, column=iter_loc + 1).value = epoch_loss_tr
            sheet.cell(row=current_row, column=iter_loc + 2).value = epoch_loss
            sheet.cell(row=current_row, column=iter_loc + 3).value = epoch_cir1_tr
            sheet.cell(row=current_row, column=iter_loc + 4).value = epoch_cir1
            book.save(logname)

        print()

    time_elapsed = time.time() - since
    print('Training complete in {:.0f}m {:.0f}s'.format(
        time_elapsed // 60, time_elapsed % 60))
    print('Best val Acc: {:4f}'.format(best_cir1))
    del inputs
    del labels
    return best_model


def train_model_both(model, criterion, optimizer, lr_scheduler, dset_loaders_real, \
                     dset_sizes_real, dset_loaders_synthetic, dset_sizes_synthetic, \
                     writer, use_gpu=True, num_epochs=25, batch_size=4, \
                     num_train=100, num_test=10, init_lr=0.001, lr_decay_epoch=7, \
                     multilabel=False, multi_prob=False, logname='logs.xlsx', iter_loc=12):
    '''
    Obsolete code. Has to be rewritten if needed
    '''
    book = openpyxl.load_workbook(logname)
    sheet = book.active
    current_row = sheet.max_row

    batch_size = batch_size * 2
    since = time.time()

    best_model = model
    best_cir1 = 0.0

    for epoch in range(num_epochs):
        print('Epoch {}/{}'.format(epoch, num_epochs - 1))
        print('-' * 10)
        # Iterate over data.
        batch_count = 0
        #
        if lr_scheduler is not None:
            optimizer = lr_scheduler(optimizer, epoch, init_lr=init_lr, lr_decay_epoch=lr_decay_epoch)
        model.train(True)  # Set model to training mode
        for opt_iter in range(num_test):
            running_loss = 0.0
            running_corrects = 0
            running_cir1 = 0

            for k in range(num_train):
                # get the inputs
                inputs_real, labels_real = next(iter(dset_loaders_real['train']))
                inputs_synthetic, labels_synthetic = next(iter(dset_loaders_synthetic['train']))
                inputs = torch.cat([inputs_real, inputs_synthetic], 0)
                labels = torch.cat([labels_real, labels_synthetic], 0)

                # wrap them in Variable
                if use_gpu:
                    inputs, labels = Variable(inputs.cuda()), \
                                     Variable(labels.cuda())
                else:
                    inputs, labels = Variable(inputs), Variable(labels)

                # zero the parameter gradients
                optimizer.zero_grad()

                # forward
                outputs = model(inputs)
                _, preds = torch.max(outputs.data, 1)
                if (not multilabel):
                    loss = criterion(outputs, labels)
                else:
                    labels_multi = []
                    for label in labels.data:
                        label_multi = np.zeros(11)

                        if (multi_prob):
                            label_multi[label] = .5
                            label_multi[label + 1] = 1
                            label_multi[label + 2] = .5
                        else:
                            label_multi[label:label + 3] = 1

                        label_multi = label_multi[1:-1]
                        labels_multi.append(label_multi)
                    labelsv = Variable(torch.FloatTensor(labels_multi).cuda()).view(-1, 9)
                    loss = criterion(outputs, labelsv)

                # backward + optimize only if in training phase
                loss.backward()
                optimizer.step()

                # statistics
                running_loss += loss.item()
                running_corrects += torch.sum(preds == labels.data)
                running_cir1 += torch.sum(torch.abs(preds - labels.data) <= 1)

            epoch_loss = running_loss / (num_train * batch_size)
            epoch_acc = running_corrects / (num_train * batch_size)
            epoch_cir1 = running_cir1 / (num_train * batch_size)
            writer.add_scalar('train loss', epoch_loss, epoch)
            writer.add_scalar('train accuracy', epoch_acc, epoch)
            writer.add_scalar('train CIR-1', epoch_cir1, epoch)

            print('{}/{}, Loss: {:.4f} Acc: {:.4f} CIR-1: {:.4f}'
                  .format(opt_iter + 1, num_test, epoch_loss, epoch_acc, epoch_cir1))

            # deep copy the model
        model.train(False)  # Set model to evaluate mode

        test_count = 0
        for loader in [dset_loaders_real, dset_loaders_synthetic]:
            running_loss = 0.0
            running_corrects = 0
            running_cir1 = 0
            test_count += 1
            for data in loader['val']:
                # get the inputs
                inputs, labels = data

                # wrap them in Variable
                if use_gpu:
                    inputs, labels = Variable(inputs.cuda()), \
                                     Variable(labels.cuda())
                else:
                    inputs, labels = Variable(inputs), Variable(labels)

                # zero the parameter gradients
                optimizer.zero_grad()

                # forward
                outputs = model(inputs)
                _, preds = torch.max(outputs.data, 1)
                if (not multilabel):
                    loss = criterion(outputs, labels)
                else:
                    labels_multi = []
                    for label in labels.data:
                        label_multi = np.zeros(11)

                        if (multi_prob):
                            label_multi[label] = .5
                            label_multi[label + 1] = 1
                            label_multi[label + 2] = .5
                        else:
                            label_multi[label:label + 3] = 1

                        label_multi = label_multi[1:-1]
                        labels_multi.append(label_multi)
                    labelsv = Variable(torch.FloatTensor(labels_multi).cuda()).view(-1, 9)
                    loss = criterion(outputs, labelsv)

                # statistics
                running_loss += loss.item()
                running_corrects += torch.sum(preds == labels.data)
                running_cir1 += torch.sum(torch.abs(preds - labels.data) <= 1)
            if (test_count == 1):
                epoch_loss = running_loss / dset_sizes_real['val']
                epoch_acc = running_corrects / dset_sizes_real['val']
                epoch_cir1 = running_cir1 / dset_sizes_real['val']
                writer.add_scalar('real val loss', epoch_loss, epoch)
                writer.add_scalar('real val accuracy', epoch_acc, epoch)
                writer.add_scalar('real val CIR-1', epoch_cir1, epoch)
                print('Real Val Loss: {:.4f} Acc: {:.4f} CIR-1: {:.4f}'.format(epoch_loss, epoch_acc, epoch_cir1))
            if (test_count == 2):
                epoch_loss = running_loss / dset_sizes_synthetic['val']
                epoch_acc = running_corrects / dset_sizes_synthetic['val']
                epoch_cir1 = running_cir1 / dset_sizes_synthetic['val']
                writer.add_scalar('synthetic val loss', epoch_loss, epoch)
                writer.add_scalar('synthetic val accuracy', epoch_acc, epoch)
                writer.add_scalar('synthetic val CIR-1', epoch_cir1, epoch)
                print('Synthetic Val Loss: {:.4f} Acc: {:.4f} CIR-1: {:.4f}'.format(epoch_loss, epoch_acc, epoch_cir1))

        if epoch_cir1 > best_cir1:
            best_cir1 = epoch_cir1
            best_model = copy.deepcopy(model)
            sheet.cell(row=current_row, column=iter_loc).value = epoch + 1
            # sheet.cell(row=current_row, column=iter_loc + 1).value = epoch_loss_tr
            sheet.cell(row=current_row, column=iter_loc + 2).value = epoch_loss
            # sheet.cell(row=current_row, column=iter_loc + 3).value = epoch_cir1_tr
            sheet.cell(row=current_row, column=iter_loc + 4).value = epoch_cir1
            book.save(logname)

        print()

    time_elapsed = time.time() - since
    print('Training complete in {:.0f}m {:.0f}s'.format(
        time_elapsed // 60, time_elapsed % 60))
    print('Best val Acc: {:4f}'.format(best_cir1))
    return model